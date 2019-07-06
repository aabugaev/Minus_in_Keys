import subprocess
import sys

def install(package):
    subprocess.call([sys.executable, "-m", "pip", "install", package])


install('pymorphy2')
install('openpyxl')

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import pymorphy2

morph = pymorphy2.MorphAnalyzer()
from itertools import chain
import re

minuswordslist = []
redFill = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')


# In[42]:


MP_name  = str(input("Please give the name of your MP.\n"))
mp_wb = load_workbook(MP_name)
#mp_wb_ws = wb.get_active_sheet()


# In[43]:


mp_sheet_with_minus_name = str(input("Please give the name of the Sheet with minuswords.\n"))
mp_wb_sheet_with_minus = mp_wb[mp_sheet_with_minus_name]

Minuswords_All_Lines = [minus_word_line  for minus_word_line in mp_wb_sheet_with_minus]
Minuswords_All_Cells = list(chain.from_iterable(Minuswords_All_Lines))


# In[44]:


for Minuswords_Cell in Minuswords_All_Cells:
    if Minuswords_Cell.value is not None:
        current_word = morph.parse(str(Minuswords_Cell.value))[0]
        for lex in current_word.lexeme:
            minuswordslist.append(str(lex.word))
print(minuswordslist)


# In[45]:


mp_sheet_with_keys_name = str(input("Please give the name of the Sheet with keywords.\n"))


# In[46]:


mp_sheet_with_keys = mp_wb[mp_sheet_with_keys_name]


Keywords_All_Lines = [keyword_line  for keyword_line in mp_sheet_with_keys]
Keywords_All_Cells = list(chain.from_iterable(Keywords_All_Lines))


# In[47]:


for Keyword_Cell in Keywords_All_Cells:
    for minusword in minuswordslist:
    	if bool(re.search(r"\b" + minusword + r"\b", str(Keyword_Cell.value), re.IGNORECASE)):
        	print(minusword + " == found in ==" + str(Keyword_Cell.value))
        	Keyword_Cell.fill = redFill           	
mp_wb.save("Minuswords_found_in_"+MP_name)


# In[48]:


"""

Без кавычек:

#%load file.py
%%writefile file.py  - в начале блока
%pycat  -
%run file.py
%lsmagic

from IPython.core.interactiveshell import InteractiveShell
InteractiveShell.ast_node_interactivity = "all"

===openpyxl===
Минус-словаСлова Директ + Adwords

wrwb = Workbook()
wrwb_ws = wrwb.get_active_sheet()

wb.save()

===numpy/pandas===
import pandas as pd
import numpy as np

excel_df = pd.read_excel()
csv_df = pd.read_csv()


df.to_excel()
df.to_csv()

writer = pd.ExcelWriter('',engine='xlsxwriter',options={})
df.to_excel(writer)
writer.save()


====requests/BeautifulSoup===
import requests
from bs4 import BeautifulSoup

page = requests.get("http://yandex.ru")
page.encoding = "windows-1251"
soup = BeautifulSoup(''.join(page.text), "html.parser\"),
soup.findAll("div")


===Files and directories===
import os
FileList = os.listdir()

#if not os.path.exists("Folder"):
#   os.mkdir("Folder") 

"""

